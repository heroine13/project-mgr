"""
数据导出工具模块
支持 Excel、CSV、PDF 格式导出
"""
import csv
import io
from typing import List, Dict, Any, Optional
from datetime import datetime
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class BaseExporter:
    """导出器基类"""
    
    def __init__(self, title: str = "Export"):
        self.title = title
        self.headers: List[str] = []
        self.data: List[List[Any]] = []
    
    def set_data(self, headers: List[str], rows: List[List[Any]]):
        """设置导出数据"""
        self.headers = headers
        self.data = rows
    
    def export(self) -> bytes:
        """导出数据，子类实现"""
        raise NotImplementedError


class ExcelExporter(BaseExporter):
    """Excel 导出器"""
    
    def __init__(self, title: str = "Export"):
        super().__init__(title)
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")
    
    def export(self) -> bytes:
        """导出为 Excel 格式"""
        wb = Workbook()
        ws = wb.active
        ws.title = self.title
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 写入数据
        for row_idx, row in enumerate(self.data, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # 自动调整列宽
        for col_idx in range(1, len(self.headers) + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class CSVExporter(BaseExporter):
    """CSV 导出器"""
    
    def export(self) -> bytes:
        """导出为 CSV 格式"""
        output = io.StringIO()
        
        # 写入 BOM (支持中文)
        output.write('\ufeff')
        
        writer = csv.writer(output)
        
        # 写入表头
        writer.writerow(self.headers)
        
        # 写入数据
        for row in self.data:
            writer.writerow([str(cell) if cell is not None else '' for cell in row])
        
        output.seek(0)
        return output.getvalue().encode('utf-8')


class PDFExporter(BaseExporter):
    """PDF 导出器"""
    
    def __init__(self, title: str = "Export"):
        super().__init__(title)
        if not PDF_AVAILABLE:
            raise ImportError("reportlab 未安装，请运行: pip install reportlab")
    
    def export(self) -> bytes:
        """导出为 PDF 格式"""
        output = io.BytesIO()
        
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # 创建故事
        story = []
        
        # 标题样式
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=getSampleStyleSheet()['Heading1'],
            fontSize=18,
            textColor=colors.HexColor("#4472C4"),
            spaceAfter=30
        )
        
        # 添加标题
        story.append(Paragraph(self.title, title_style))
        story.append(Spacer(1, 12))
        
        # 添加导出时间
        time_style = ParagraphStyle(
            'TimeStyle',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=10,
            textColor=colors.gray
        )
        story.append(Paragraph(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", time_style))
        story.append(Spacer(1, 20))
        
        # 创建表格数据
        table_data = [self.headers]
        table_data.extend(self.data)
        
        # 创建表格
        table = Table(table_data)
        
        # 表格样式
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # 交替行颜色
        for i in range(1, len(table_data)):
            if i % 2 == 0:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.white)
            else:
                table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F5F5F5"))
        
        table.setStyle(table_style)
        story.append(table)
        
        # 构建 PDF
        doc.build(story)
        output.seek(0)
        return output.getvalue()


# === 便捷导出函数 ===

def export_to_excel(headers: List[str], rows: List[List[Any]], filename: str = "export.xlsx") -> bytes:
    """导出为 Excel"""
    exporter = ExcelExporter(filename.replace('.xlsx', ''))
    exporter.set_data(headers, rows)
    return exporter.export()


def export_to_csv(headers: List[str], rows: List[List[Any]], filename: str = "export.csv") -> bytes:
    """导出为 CSV"""
    exporter = CSVExporter(filename.replace('.csv', ''))
    exporter.set_data(headers, rows)
    return exporter.export()


def export_to_pdf(headers: List[str], rows: List[List[Any]], filename: str = "export.pdf") -> bytes:
    """导出为 PDF"""
    exporter = PDFExporter(filename.replace('.pdf', ''))
    exporter.set_data(headers, rows)
    return exporter.export()


# === 数据格式化工具 ===

def format_task_for_export(tasks: List[Dict[str, Any]]) -> tuple:
    """格式化任务数据用于导出"""
    headers = ['ID', '任务名称', '项目', '负责人', '状态', '优先级', '开始日期', '截止日期', '进度%', '描述']
    rows = []
    
    for task in tasks:
        row = [
            task.get('id', ''),
            task.get('title', ''),
            task.get('project_name', ''),
            task.get('assignee_name', ''),
            task.get('status', ''),
            task.get('priority', ''),
            task.get('start_date', ''),
            task.get('due_date', ''),
            task.get('progress', 0),
            task.get('description', '')[:100]  # 截断长描述
        ]
        rows.append(row)
    
    return headers, rows


def format_project_for_export(projects: List[Dict[str, Any]]) -> tuple:
    """格式化项目数据用于导出"""
    headers = ['ID', '项目名称', '负责人', '状态', '开始日期', '结束日期', '预算', '进度%', '描述']
    rows = []
    
    for project in projects:
        row = [
            project.get('id', ''),
            project.get('name', ''),
            project.get('owner_name', ''),
            project.get('status', ''),
            project.get('start_date', ''),
            project.get('end_date', ''),
            project.get('budget', 0),
            project.get('progress', 0),
            project.get('description', '')[:100]
        ]
        rows.append(row)
    
    return headers, rows


def format_user_for_export(users: List[Dict[str, Any]]) -> tuple:
    """格式化用户数据用于导出"""
    headers = ['ID', '用户名', '邮箱', '全名', '状态', '创建时间']
    rows = []
    
    for user in users:
        row = [
            user.get('id', ''),
            user.get('username', ''),
            user.get('email', ''),
            user.get('full_name', ''),
            '激活' if user.get('is_active') else '禁用',
            user.get('created_at', '')[:10]
        ]
        rows.append(row)
    
    return headers, rows